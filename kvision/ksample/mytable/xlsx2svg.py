#encoding=utf8
import re, os, sys
import cv2
from PIL import Image

CURFILEDIR = os.path.split(os.path.abspath(__file__))[0]
PYTHONXDIR = r"E:\kSource"
if not sys.path or PYTHONXDIR not in sys.path: # for pythonx
    sys.path.insert(0, PYTHONXDIR)
    sys.path.insert(0, os.path.join(PYTHONXDIR, "pythonx"))
from pythonx.funclib import *

import tarfile
import time
from functools import wraps
from copy import deepcopy

# Use Aspose.Cells for Python via Java
import jpype
import asposecells
jpype.startJVM()
from asposecells.api import *

# https://blog.aspose.com/cells/convert-excel-to-image-in-python/#Convert-Excel-to-SVG-in-Python
def xlsx2svgf(xlsxfile, svgfile, svg=True):
    if os.path.exists(svgfile):
        return

    # load the Excel workbook
    workbook = Workbook(xlsxfile) # 这里会崩溃，说是超过了评估版本打开的限制。

    # create image options
    # https://gist.github.com/aspose-com-gists/5d33fa768a61d24704a7350432266781
    imgOptions = ImageOrPrintOptions()
    if svg:
        imgOptions.setSaveFormat(SaveFormat.SVG)
    else:
        imgOptions.setSaveFormat(SaveFormat.PNG)

    # get sheet count
    sheetCount = workbook.getWorksheets().getCount()

    # loop through the sheets
    sheet = workbook.getWorksheets().get(0)

    # convert each sheet to SVG
    sr = SheetRender(sheet, imgOptions)
    #for j in range(0, sr.getPageCount()):
    try:
        sr.toImage(0, svgfile)
    except java.lang.IndexOutOfBoundsException: # java.lang.IndexOutOfBoundsException: Index: 0, Size: 0
        pass
    workbook.dispose()

def formatxls(fpath, force=False):

    if fpath.endswith(".xlsx.xlsx"):
        return

    fmd5 = getmd5(os.path.abspath(fpath))
    #if not force and getFileMd5(fpath) == readfile(os.path.join("tempdir", "xlsx", "result", fmd5[:8], fmd5[8:]), True):
    #    return

    import openpyxl
    from openpyxl.styles import Side, Border

    print("formatxls", fpath)

    # 加载Excel文件
    wb = openpyxl.load_workbook(fpath)

    # 设置线条的样式和颜色
    side = Side(style="thin", color="101010") # thick
    # 设置单元格的边框线条
    border = Border(top=side, bottom=side, left=side, right=side)

    # 获取最大行数
    #max_row = sheet.max_row
    # 获取最大列数
    #max_column = sheet.max_column

    sheet = wb.active
    rows = sheet.rows
    # ws.cell(row=0, column=0).border = border
    for row in rows:
        for cell in row:
            #  'bottom', 'copy', 'diagonal', 'diagonalDown', 'diagonalUp', 'diagonal_direction', 'end',
            # 'horizontal', 'left', 'outline', 'right', 'start', 'top', 'vertical']
            #print(dir(cell.border))
            # 'border_style', 'color', 'from_tree', 'idx_base', 'namespace', 'style', 'tagname', 'to_tree']
            #print(dir(cell.border.top))
            #print(cell.border.top.color)
            #print(cell.border.top.color.rgb)
            if cell.border and cell.border.top and cell.border.top.color and cell.border.top.color.rgb:
                if "00101010" == cell.border.top.color.rgb:
                    #print(type(cell.border.top), cell.border.top)
                    #print(dir(cell.border.top))
                    wb.close()

                    fmd5 = getmd5(os.path.abspath(fpath))
                    #writefile(os.path.join("tempdir", "xlsx", "result", fmd5[:8], fmd5[8:]), getFileMd5(fpath))
                    return

            cell.border = border

    print("formatxls", fpath)

    # 处理完成后保存表格，会在当前目录生成一个excel文件
    wb.save(filename=fpath) # +".xlsx"
    # 关闭表格对象
    wb.close()

    fmd5 = getmd5(os.path.abspath(fpath))
    #writefile(os.path.join("tempdir", "xlsx", "result", fmd5[:8], fmd5[8:]), getFileMd5(fpath))

def myxlsx2svg(rootdir):

    def mainfile(fpath, fname, ftype):
        if not ftype in ("xlsx"):
            return
        print(fpath)
        formatxls(fpath)
        xlsx2svgf(fpath, fpath+".svg", True)
        #xlsx2svgf(fpath, fpath+".png", False)

    searchdir(rootdir, mainfile)

if __name__ == "__main__":
    if not "test" in sys.argv:
        rootdir = r"E:\kSource\blog\kvision\ksample\mytable"
        myxlsx2svg(rootdir)
        rootdir = r"E:\kSource\blog\kvision\ksample\imgtable_v3"
        myxlsx2svg(rootdir)
        print("ok")
    else:
        formatxls(r"E:\kSource\blog\kvision\ksample\mytable\quark\001_8e68c.xlsx", True)
        formatxls(r"E:\kSource\blog\kvision\ksample\mytable\BaiduOCRConverter_Excel\001_8e68c.xlsx", True)
        print("ok")
